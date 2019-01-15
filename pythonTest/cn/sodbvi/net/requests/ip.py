import json

import requests
from openpyxl import load_workbook


def readFile():
    phoneDict = dict()
    workbook = load_workbook('D:\\tj.xlsx')
    sheets = workbook.get_sheet_names()  # 从名称获取sheet
    booksheet = workbook.get_sheet_by_name(sheets[0])

    rows = booksheet.rows
    # 迭代所有的行
    for row in rows:
        for ceil in row:
            phoneDict[ceil.value] = ''

    return phoneDict


def writeFile(phoneDict):
    workbook = load_workbook('D:\\tj.xlsx')
    sheets = workbook.get_sheet_names()  # 从名称获取sheet
    booksheet = workbook.get_sheet_by_name(sheets[0])

    rows = booksheet.rows

    workbook.save("'D:\\tj.xlsx")


def tj(pDict):
    headers = {'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8',
               'Accept-Encoding': 'gzip, deflate, compress',
               'Accept-Language': 'en-us;q=0.5,en;q=0.3',
               'Cache-Control': 'max-age=0',
               'Connection': 'keep-alive',
               'User-Agent': 'Mozilla/5.0 (X11; Ubuntu; Linux x86_64; rv:22.0) Gecko/20100101 Firefox/22.0'}

    s = requests.Session()
    s.headers.update(headers)
    for k in pDict:
        url = 'http://mobsec-dianhua.baidu.com/dianhua_api/open/location?tel=' + str(k)
        r = s.get(url)
        pDict[k] = json.loads(r.text)['response'][str(k)]['detail']['province']
        print(str(k) + ',' + pDict[k])
    return pDict


if __name__ == '__main__':
    pDict = readFile()
    writeFile(tj(pDict))
